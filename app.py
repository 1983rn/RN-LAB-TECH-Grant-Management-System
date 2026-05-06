#!/usr/bin/env python3
"""
Grant Management System - Web Application
Multi-Tenant with Mandatory Authentication
"""

import sys
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, OSError, ValueError):
        pass

import json
import os
from datetime import datetime, timedelta
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, session, flash
from flask_cors import CORS
from io import BytesIO

# Import authentication modules
from database import init_database, get_db, log_action, generate_otp, hash_password, create_developer_account, DATABASE_PATH as DB_FILE
from auth import login_school, login_developer, require_login, require_developer, get_current_school_id
import uuid
from subscription_plans import SUBSCRIPTION_PLANS, get_plan

# Import database helpers for multi-tenant data access
from db_helpers import (
    get_school_settings, save_school_settings,
    get_school_budget, save_school_budget,
    get_school_credits, save_school_credit, delete_school_credit,
    get_school_debits, save_school_debit, delete_school_debit,
    generate_transaction_number, initialize_budget_for_year,
    generate_ipdc_number, update_debit_ipdc_fields,
    get_term_settings, save_term_settings
)

# Try to import openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    EXCEL_AVAILABLE = True
    print("openpyxl loaded successfully - Excel export enabled")
except ImportError as e:
    EXCEL_AVAILABLE = False
    print(f"⚠️  WARNING: openpyxl import failed: {e}")
    print("   Install with: pip install openpyxl==3.1.2")

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get('SECRET_KEY', 'rn-lab-tech-grant-mgmt-secret-2025-stable-key')

# In-memory cache for IPDC sessions
ipdc_sessions = {}

# Custom filters
@app.template_filter('from_json')
def from_json_filter(value):
    if not value:
        return {}
    try:
        return json.loads(value)
    except:
        return {}

print(f"Database location: {DB_FILE}")

# Initialize database on startup
with app.app_context():
    init_database()
    create_developer_account()
    print("✅ Database and developer account initialized")

# Track school activity
@app.before_request
def update_last_activity():
    """Update last_activity timestamp for logged-in schools"""
    if 'user' in session and not session.get('user', {}).get('is_developer'):
        school_id = get_current_school_id()
        if school_id:
            try:
                timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
                with get_db() as conn:
                    cursor = conn.cursor()
                    cursor.execute('''
                        UPDATE schools 
                        SET last_activity = ? 
                        WHERE id = ?
                    ''', (timestamp, school_id))
                print(f"✓ Updated activity for school_id={school_id} at {timestamp}")
            except Exception as e:
                print(f"✗ Error updating activity for school_id={school_id}: {e}")

# ============================================================================
# SUBSCRIPTION MANAGEMENT FUNCTIONS
# ============================================================================

def calculate_days_remaining(expiry_date_str):
    """Calculate days remaining until subscription expires"""
    if not expiry_date_str or expiry_date_str == 'None':
        return 0
    try:
        today = datetime.utcnow().date()
        expiry = datetime.strptime(str(expiry_date_str), "%Y-%m-%d").date()
        return (expiry - today).days
    except Exception as e:
        print(f"Error calculating days remaining: {e}")
        return 0

def get_subscription_info(school_id):
    """Get subscription information for a school"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT subscription_type, expiry_date, subscription_end
                FROM schools WHERE id = ?
            ''', (school_id,))
            school = cursor.fetchone()
            if not school:
                return None
            
            expiry_date = school['expiry_date'] or school['subscription_end'] or None
            days_remaining = calculate_days_remaining(expiry_date) if expiry_date else 0
            
            if days_remaining > 14:
                status_class = 'subscription-active'
            elif 0 < days_remaining <= 14:
                status_class = 'subscription-warning'
            else:
                status_class = 'subscription-expired'
            
            return {
                'subscription_type': school['subscription_type'] or 'TRIAL',
                'expiry_date': expiry_date or 'Not Set',
                'days_remaining': days_remaining,
                'status_class': status_class
            }
    except Exception as e:
        print(f"Error getting subscription info: {e}")
        return None

def send_warning_message(school):
    """Send expiry warning message to school"""
    message = f"""Dear {school['school_name']},

Your subscription will expire in 14 days on {school['expiry_date']}.
Please renew to avoid service interruption.

Contact Developer:
Phone: +265991332952
WhatsApp: +265999630132
Email: robertnsambe@gmail.com"""
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO subscription_messages
            (school_id, message, message_type)
            VALUES (?, ?, ?)
        ''', (school['id'], message, 'WARNING'))

def auto_expiry_check():
    """Check for expiring subscriptions and send warnings"""
    target_date = (datetime.utcnow().date() + timedelta(days=14)).strftime('%Y-%m-%d')
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, school_name, username, expiry_date, subscription_end
            FROM schools
            WHERE school_name != 'DEVELOPER_ACCOUNT'
        ''')
        schools = cursor.fetchall()
        
        warned_count = 0
        for school in schools:
            expiry = school['expiry_date'] or school['subscription_end']
            if expiry and expiry == target_date:
                send_warning_message(dict(school))
                warned_count += 1
        
        return warned_count

# ============================================================================
# MULTI-TENANT DATABASE FUNCTIONS (Replaces JSON file operations)
# ============================================================================

def get_budget(academic_year, term):
    """Get budget for specific school, academic year and term"""
    school_id = get_current_school_id()
    if not school_id:
        return None
    budget = get_school_budget(school_id, academic_year, term)
    
    # Get settings for totalGrant and schoolName
    settings = get_school_settings(school_id, academic_year, term)
    
    if budget:
        # Calculate total available funds (New Grant + Balance B/F)
        total = float(settings.get('total_grant', 0)) + float(settings.get('balance_bf', 0))
        # Add fields for template compatibility
        budget['totalGrant'] = total
        budget['total_grant'] = total
        budget['schoolName'] = session.get('school_name', '')
        return budget
    
    # Calculate total available funds (New Grant + Balance B/F)
    total = float(settings.get('total_grant', 0)) + float(settings.get('balance_bf', 0))
    # Return empty structure if no budget exists
    return {
        'academicYear': academic_year,
        'term': term,
        'totalGrant': total,
        'total_grant': total,
        'schoolName': session.get('school_name', ''),
        'items': []
    }

def save_budget(budget_data):
    """Save budget data for specific school"""
    school_id = get_current_school_id()
    if not school_id:
        return False
    academic_year = budget_data.get('academicYear', get_academic_year())
    term = budget_data.get('term', get_term())
    return save_school_budget(school_id, academic_year, term, budget_data)

def get_credits(academic_year, term):
    """Get credits for specific school, academic year and term"""
    school_id = get_current_school_id()
    if not school_id:
        return []
    return get_school_credits(school_id, academic_year, term)

def get_settings():
    """Get system settings for current school, academic year and term"""
    school_id = get_current_school_id()
    if not school_id:
        return {
            'compiledBy': '',
            'enteredBy': '',
            'authorizingOfficer': '',
            'authorizingAppointment': '',
            'counterSign': '',
            'counterAppointment': '',
            'ministry': 'Education',
            'academicYear': '2026-2027',
            'term': 'Term 1',
            'schoolName': '',
            'schoolAddress': '',
            'totalGrant': 0,
            'total_grant': 0,
            'balanceBF': 0,
            'balance_bf': 0
        }
    
    academic_year = get_academic_year()
    term = get_term()
    
    db_settings = get_school_settings(school_id, academic_year, term)
    
    # Convert database format to app format with both naming conventions
    return {
        'compiledBy': db_settings.get('compiled_by', ''),
        'enteredBy': db_settings.get('entered_by', ''),
        'authorizingOfficer': db_settings.get('authorizing_officer', ''),
        'authorizingAppointment': db_settings.get('authorizing_appointment', ''),
        'counterSign': db_settings.get('counter_sign', ''),
        'counterAppointment': db_settings.get('counter_appointment', ''),
        'ministry': db_settings.get('ministry_department', 'Education'),
        'academicYear': academic_year,
        'term': term,
        'schoolName': db_settings.get('school_name', ''),
        'schoolAddress': db_settings.get('school_address', ''),
        'totalGrant': db_settings.get('total_grant', 0),
        'total_grant': db_settings.get('total_grant', 0),
        'balanceBF': db_settings.get('balance_bf', 0),
        'balance_bf': db_settings.get('balance_bf', 0)
    }

def get_academic_year():
    """Get current academic year from session or term settings"""
    if 'academic_year' in session:
        return session['academic_year']
    
    school_id = get_current_school_id()
    if school_id:
        settings = get_term_settings(school_id)
        session['academic_year'] = settings['current_academic_year']
        return settings['current_academic_year']
    
    return '2026-2027'

def get_term():
    """Get current term from session or term settings"""
    if 'term' in session:
        return session['term']
    
    school_id = get_current_school_id()
    if school_id:
        settings = get_term_settings(school_id)
        session['term'] = settings['current_term']
        return settings['current_term']
    
    return 'Term 1'

def get_financial_year():
    """Compatibility: returns academic year"""
    return get_academic_year()

def get_total_grant(academic_year=None, term=None):
    """Get total grant for specific school and academic year/term"""
    school_id = get_current_school_id()
    if not school_id:
        return 0
    if not academic_year:
        academic_year = get_academic_year()
    if not term:
        term = get_term()
    settings = get_school_settings(school_id, academic_year, term)
    total = float(settings.get('total_grant', 0)) + float(settings.get('balance_bf', 0))
    return total

def save_settings(settings_data):
    """Save settings for current school, academic year and term"""
    school_id = get_current_school_id()
    if not school_id:
        return False
    academic_year = settings_data.get('academicYear', get_academic_year())
    term = settings_data.get('term', get_term())
    return save_school_settings(school_id, academic_year, term, settings_data)

def save_credit(credit_data):
    """Save credit for current school"""
    school_id = get_current_school_id()
    if not school_id:
        return False
    academic_year = credit_data.get('academicYear', get_academic_year())
    term = credit_data.get('term', get_term())
    return save_school_credit(school_id, academic_year, term, credit_data)

def get_debits(academic_year, term):
    """Get debits for specific school, academic year and term"""
    school_id = get_current_school_id()
    if not school_id:
        return []
    return get_school_debits(school_id, academic_year, term)

def save_debit(debit_data):
    """Save debit for current school"""
    school_id = get_current_school_id()
    if not school_id:
        return False
    academic_year = debit_data.get('academicYear', get_academic_year())
    term = debit_data.get('term', get_term())
    debit_id = save_school_debit(school_id, academic_year, term, debit_data)
    if debit_id:
        debit_data['id'] = f"debit_{debit_id}"
        return True
    return False

def calculate_spending(budget, debits):
    """Calculate spending for budget items"""
    if not budget:
        return []
    
    items = []
    for budget_item in budget.get('items', []):
        item_id = budget_item.get('id')
        item_debits = [d for d in debits if d.get('itemId') == item_id]
        spent = sum(d.get('amount', 0) for d in item_debits)
        
        items.append({
            'id': item_id,
            'powNo': budget_item.get('powNo'),
            'powName': budget_item.get('powName'),
            'subActivity': budget_item.get('subActivity'),
            'subItemDescription': budget_item.get('subItemDescription'),
            'code': budget_item.get('code'),
            'totalAllocation': budget_item.get('totalAllocation', 0),
            'spent': spent,
            'balance': budget_item.get('totalAllocation', 0) - spent
        })
    
    return items

def generate_budget_structure():
    """Master template for Malawi Grant Management - 43 rows across 16 POWs"""
    months = ["April", "May", "June", "July", "August", "September", "October", "November", "December", "January", "February", "March"]
    empty_monthly = {month: 0 for month in months}
    
    # Master template - each row is unique by template_row_id
    template = [
        # POW 1 - Facilitating office operations (14 rows)
        {'template_row_id': 1, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Wages for support staff', 'code': '2211012204'},
        {'template_row_id': 2, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Public transport', 'code': '2211011203'},
        {'template_row_id': 3, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Heating and lighting', 'code': '2211011401'},
        {'template_row_id': 4, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Telephone charges', 'code': '2211011402'},
        {'template_row_id': 5, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Water and sanitation', 'code': '2211011405'},
        {'template_row_id': 6, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Consumable stores', 'code': '2211011502'},
        {'template_row_id': 7, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Postage', 'code': '2211011504'},
        {'template_row_id': 8, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Printing cost', 'code': '2211011505'},
        {'template_row_id': 9, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Publication and advertisement', 'code': '2211011406'},
        {'template_row_id': 10, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Stationery', 'code': '2211011506'},
        {'template_row_id': 11, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Uniform and protective wear', 'code': '2211011507'},
        {'template_row_id': 12, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Fuel and Lubricants', 'code': '2211012401'},
        {'template_row_id': 13, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Subscriptions', 'code': '2211012321'},
        {'template_row_id': 14, 'powNo': '1', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Facilitating office operations', 'subItemDescription': 'Purchase of plant and office equipment', 'code': '2211010251'},
        # POW 2 (1 row)
        {'template_row_id': 15, 'powNo': '2', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Management of school based and National examinations', 'subItemDescription': 'Examinations', 'code': '2211011803'},
        # POW 3 (2 rows)
        {'template_row_id': 16, 'powNo': '3', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Conducting budget and management meetings', 'subItemDescription': 'Fuel 0r 2103 public transport', 'code': '2211012401'},
        {'template_row_id': 17, 'powNo': '3', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Conducting budget and management meetings', 'subItemDescription': 'Subsistence allowance', 'code': '2211011204'},
        # POW 4 (2 rows)
        {'template_row_id': 18, 'powNo': '4', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'SMASSE', 'subItemDescription': 'Fuel or 2103 public transport', 'code': '2211012401'},
        {'template_row_id': 19, 'powNo': '4', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'SMASSE', 'subItemDescription': 'Subsistence allowance', 'code': '2211011204'},
        # POW 5 (3 rows)
        {'template_row_id': 20, 'powNo': '5', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Sporting activities', 'subItemDescription': 'Sporting equipment', 'code': '2211011805'},
        {'template_row_id': 21, 'powNo': '5', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Sporting activities', 'subItemDescription': 'Fuel or 1203-Public transport', 'code': '2211012401'},
        {'template_row_id': 22, 'powNo': '5', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Sporting activities', 'subItemDescription': 'Subsistence allowance', 'code': '2211011204'},
        # POW 6 (1 row)
        {'template_row_id': 23, 'powNo': '6', 'powName': 'Improve access for special needs education', 'subActivity': 'Support to SNE', 'subItemDescription': 'Purchase of special needs materials', 'code': '2211011806'},
        # POW 7 (3 rows)
        {'template_row_id': 24, 'powNo': '7', 'powName': 'Procure and supply Secondary TLMs', 'subActivity': 'Procurement of teaching and learning materials', 'subItemDescription': 'Science consumables', 'code': '2211011807'},
        {'template_row_id': 25, 'powNo': '7', 'powName': 'Procure and supply Secondary TLMs', 'subActivity': 'Procurement of teaching and learning materials', 'subItemDescription': 'Text books', 'code': '2211011804'},
        {'template_row_id': 26, 'powNo': '7', 'powName': 'Procure and supply Secondary TLMs', 'subActivity': 'Procurement of teaching and learning materials', 'subItemDescription': 'Purchase of school supplies', 'code': '2211011808'},
        # POW 8 (2 rows)
        {'template_row_id': 27, 'powNo': '8', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Support to HIV/AIDS related activities', 'subItemDescription': 'HIV/AIDS services', 'code': '2211011614'},
        {'template_row_id': 28, 'powNo': '8', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Support to learners with first aid.', 'subItemDescription': 'Drugs', 'code': '2211011601'},
        # POW 9 (2 rows)
        {'template_row_id': 29, 'powNo': '9', 'powName': 'Maintenance of infrastructure in secondary schools', 'subActivity': 'Maintenance of infrastructure', 'subItemDescription': 'Maintenance of buildings', 'code': '2211012501'},
        {'template_row_id': 30, 'powNo': '9', 'powName': 'Maintenance of infrastructure in secondary schools', 'subActivity': 'Maintenance of infrastructure', 'subItemDescription': 'Maintenance of water supplies', 'code': '2211012504'},
        # POW 10 (2 rows) - Fixed: Only COSOMA and Computer Service
        {'template_row_id': 31, 'powNo': '10', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'COSOMA', 'subItemDescription': 'Subscription', 'code': '2211012321'},
        {'template_row_id': 32, 'powNo': '10', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Computer Service Subscription', 'subItemDescription': 'Subscription', 'code': '2211012321'},
        # POW 11 (3 rows)
        {'template_row_id': 33, 'powNo': '11', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'In-service training for teachers', 'subItemDescription': 'Consumables', 'code': '2211011502'},
        {'template_row_id': 34, 'powNo': '11', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'In-service training for teachers', 'subItemDescription': 'Subsistence Allowances', 'code': '2211011204'},
        {'template_row_id': 35, 'powNo': '11', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'In-service training for teachers', 'subItemDescription': 'Public transport or 2401 fuel', 'code': '2211011203'},
        # POW 12 (2 rows)
        {'template_row_id': 36, 'powNo': '12', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Processing of Payment vouchers', 'subItemDescription': 'Subsistence Allowances', 'code': '2211011204'},
        {'template_row_id': 37, 'powNo': '12', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Processing of Payment vouchers', 'subItemDescription': 'Public transport or 2401 fuel', 'code': '2211011203'},
        # POW 13 (1 row)
        {'template_row_id': 38, 'powNo': '13', 'powName': 'Provision of sanitary pads to girls in secondary schools', 'subActivity': 'Provision of sanitary pads to girls in secondary schools', 'subItemDescription': 'Consumables', 'code': '2211011502'},
        # POW 14 (1 row)
        {'template_row_id': 39, 'powNo': '14', 'powName': 'Provision of PPEs to schools', 'subActivity': 'Provision of PPEs to schools', 'subItemDescription': 'Consumables', 'code': '2211011502'},
        # POW 15 (1 row)
        {'template_row_id': 40, 'powNo': '15', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Provision of food and other boarding necessities to learners', 'subItemDescription': 'Boarding expenses', 'code': '2211011801'},
        # POW 16 (2 rows)
        {'template_row_id': 41, 'powNo': '16', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Education visits', 'subItemDescription': 'Subsistence Allowances', 'code': '2211011204'},
        {'template_row_id': 42, 'powNo': '16', 'powName': 'Improving teaching and learning in schools', 'subActivity': 'Education visits', 'subItemDescription': 'Public transport or 2401 fuel', 'code': '2211011203'},
    ]
    
    # Add common fields to each row
    for row in template:
        row['id'] = f"pow{row['powNo']}_row{row['template_row_id']}"
        row['totalAllocation'] = 0
        row['monthlyAllocations'] = empty_monthly.copy()
    
    return template

def get_available_funds(academic_year, term):
    """Calculate available funds per budget item (Credits - Debits)"""
    budget = get_budget(academic_year, term)
    credits = get_credits(academic_year, term)
    debits = get_debits(academic_year, term)
    
    available = {}
    if budget:
        for item in budget.get('items', []):
            item_id = item.get('id')
            # Sum credits for this item using a generator expression
            item_credits = sum(
                float(li.get('amount', 0))
                for c in credits
                for li in c.get('lineItems', [])
                if li.get('itemId') == item_id
            )
            
            # Sum debits for this item
            item_debits = sum(float(d.get('amount', 0)) for d in debits if d.get('itemId') == item_id)
            
            available[item_id] = {
                'budgeted': item.get('totalAllocation', 0),
                'credited': item_credits,
                'spent': item_debits,
                'balance': item_credits - item_debits
            }
    return available


def get_developer_stats():
    """Get developer dashboard statistics with safe defaults and subscription info"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute("SELECT COUNT(*) as total FROM schools WHERE school_name != 'DEVELOPER_ACCOUNT'")
            total_schools = cursor.fetchone()['total']
            cursor.execute("SELECT COUNT(*) as active FROM schools WHERE is_active = 1 AND school_name != 'DEVELOPER_ACCOUNT'")
            active_schools = cursor.fetchone()['active']
            
            # Count online schools (active within last 5 minutes)
            five_min_ago = (datetime.utcnow() - timedelta(minutes=5)).strftime('%Y-%m-%d %H:%M:%S')
            cursor.execute('''
                SELECT COUNT(*) as online FROM schools 
                WHERE last_activity > ? AND school_name != 'DEVELOPER_ACCOUNT'
            ''', (five_min_ago,))
            online_schools = cursor.fetchone()['online']
            
            cursor.execute("SELECT COUNT(*) as expired FROM schools WHERE subscription_status = 'EXPIRED'")
            expired_schools = cursor.fetchone()['expired']
            
            # Get schools expiring in 14 days
            target_date = (datetime.utcnow().date() + timedelta(days=14)).strftime('%Y-%m-%d')
            cursor.execute('''
                SELECT COUNT(*) as expiring FROM schools 
                WHERE (expiry_date = ? OR subscription_end = ?) 
                AND school_name != 'DEVELOPER_ACCOUNT'
            ''', (target_date, target_date))
            expiring_soon = cursor.fetchone()['expiring']
            
            return {
                'total_schools': total_schools, 
                'active_schools': active_schools, 
                'online_schools': online_schools, 
                'expired_schools': expired_schools,
                'expiring_soon': expiring_soon
            }
    except Exception as e:
        print(f"Error getting stats: {e}")
        return {'total_schools': 0, 'active_schools': 0, 'online_schools': 0, 'expired_schools': 0, 'expiring_soon': 0}

# Routes
@app.route('/')
def index():
    """Root route - redirect to login"""
    if 'user' in session:
        if session['user'].get('is_developer'):
            return redirect(url_for('dev_dashboard'))
        return redirect(url_for('home'))
    return redirect(url_for('login_page'))

@app.route('/home')
@require_login
def home():
    """Protected home page - shows after login"""
    academic_year = get_academic_year()
    term = get_term()
    settings = get_settings()
    total_grant = get_total_grant(academic_year, term)
    
    return render_template('home.html',
                          school_name=settings.get('schoolName', session.get('school_name', '')),
                          academic_year=academic_year,
                          term=term,
                          total_grant=total_grant,
                          settings=settings)

@app.route('/login', methods=['GET', 'POST'])
def login_page():
    """Login page and authentication"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        dev_mode = request.form.get('dev_mode') == '1'
        
        if dev_mode:
            user, error = login_developer(username, password)
            if error:
                return render_template('login.html', error=error)
            session['user'] = user
            return redirect(url_for('dev_dashboard'))
        else:
            user, error = login_school(username, password)
            if error:
                return render_template('login.html', error=error)
            session['user'] = user
            
            # Check if password change required
            if user.get('must_change_password'):
                return redirect(url_for('request_otp'))
            
            # Update last_activity on login
            school_id = user.get('id')
            if school_id:
                try:
                    with get_db() as conn:
                        cursor = conn.cursor()
                        cursor.execute('''
                            UPDATE schools 
                            SET last_activity = ?, last_login = ?
                            WHERE id = ?
                        ''', (datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), 
                              datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'), 
                              school_id))
                except Exception as e:
                    print(f"Error updating login activity: {e}")
            
            return redirect(url_for('home'))
    
    return render_template('login.html')

@app.route('/request-otp')
@app.route('/password-reset')
@require_login
def request_otp():
    """Generate and send OTP for password change"""
    school_id = get_current_school_id()
    if not school_id:
        return redirect(url_for('login_page'))
    
    otp = generate_otp()
    hashed_otp = hash_password(otp)
    expiry = datetime.utcnow() + timedelta(minutes=10)
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE schools
            SET otp_code = ?, otp_expiry = ?
            WHERE id = ?
        ''', (hashed_otp, expiry.strftime('%Y-%m-%d %H:%M:%S'), school_id))
    
    print(f"\n🔐 OTP for school_id={school_id}: {otp}")
    print(f"   Expires at: {expiry.strftime('%Y-%m-%d %H:%M:%S')}\n")
    
    return render_template('verify_otp.html', otp=otp)

@app.route('/verify-otp', methods=['GET', 'POST'])
@require_login
def verify_otp():
    """Verify OTP and allow password change"""
    if request.method == 'GET':
        return render_template('verify_otp.html')
    
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    entered_otp = request.form.get('otp')
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT otp_code, otp_expiry
            FROM schools
            WHERE id = ?
        ''', (school_id,))
        school = cursor.fetchone()
    
    if not school or not school['otp_code']:
        return render_template('verify_otp.html', error='No OTP found. Please request a new one.')
    
    if datetime.utcnow() > datetime.strptime(school['otp_expiry'], '%Y-%m-%d %H:%M:%S'):
        return render_template('verify_otp.html', error='OTP expired. Please request a new one.')
    
    hashed_entered = hash_password(entered_otp)
    if hashed_entered != school['otp_code']:
        return render_template('verify_otp.html', error='Invalid OTP. Please try again.')
    
    return render_template('create_new_password.html')

@app.route('/create-new-password', methods=['GET', 'POST'])
@require_login
def create_new_password():
    """Create new password after OTP verification"""
    if request.method == 'GET':
        return render_template('create_new_password.html')
    
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    new_password = request.form.get('password', '').strip()
    confirm_password = request.form.get('confirm_password', '').strip()
    
    if new_password != confirm_password:
        return render_template('create_new_password.html', error='Passwords do not match')
    
    if len(new_password) < 6:
        return render_template('create_new_password.html', error='Password must be at least 6 characters')
    
    hashed_password = hash_password(new_password)
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE schools
            SET password_hash = ?,
                must_change_password = 0,
                otp_code = NULL,
                otp_expiry = NULL
            WHERE id = ?
        ''', (hashed_password, school_id))
    
    log_action('SCHOOL', str(school_id), 'PASSWORD_CHANGED', school_id, request.remote_addr)
    
    return redirect(url_for('home'))

@app.route('/logout')
def logout():
    """Logout and clear session"""
    if 'user' in session:
        school_id = get_current_school_id()
        if school_id:
            with get_db() as conn:
                cursor = conn.cursor()
                cursor.execute('UPDATE school_sessions SET is_online = 0 WHERE school_id = ? AND is_online = 1', (school_id,))
        session.pop('user', None)
    return redirect(url_for('login_page'))

# Developer Dashboard Routes
@app.route('/dev/dashboard')
@require_developer
def dev_dashboard():
    """Developer dashboard with subscription info and online status"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, school_name, username, is_active, is_locked, 
                   subscription_type, subscription_status, expiry_date, subscription_end, 
                   last_login, last_activity
            FROM schools WHERE school_name != 'DEVELOPER_ACCOUNT'
            ORDER BY school_name
        ''')
        schools_raw = cursor.fetchall()
        
        # Add days_remaining and is_online to each school
        schools = []
        current_time = datetime.utcnow()
        print(f"\n=== ONLINE STATUS DEBUG ===")
        print(f"Current UTC time: {current_time.strftime('%Y-%m-%d %H:%M:%S')}")
        
        for school in schools_raw:
            school_dict = dict(school)
            
            # Calculate days remaining
            expiry = school_dict.get('expiry_date') or school_dict.get('subscription_end') or None
            school_dict['days_remaining'] = calculate_days_remaining(expiry) if expiry else 0
            
            # Calculate online status (active within last 5 minutes)
            is_online = False
            last_activity = school_dict.get('last_activity')
            
            print(f"\nSchool: {school_dict['school_name']}")
            print(f"  last_activity: {last_activity}")
            
            if last_activity:
                try:
                    last_time = datetime.strptime(str(last_activity), '%Y-%m-%d %H:%M:%S')
                    time_diff = current_time - last_time
                    seconds_ago = time_diff.total_seconds()
                    minutes_ago = seconds_ago / 60
                    
                    print(f"  Minutes ago: {minutes_ago:.2f}")
                    print(f"  Seconds ago: {seconds_ago:.2f}")
                    
                    if seconds_ago < 300:  # 5 minutes = 300 seconds
                        is_online = True
                        print(f"  Status: ONLINE ✓")
                    else:
                        print(f"  Status: OFFLINE (too old)")
                except Exception as e:
                    print(f"  Error parsing: {e}")
            else:
                print(f"  Status: OFFLINE (no activity)")
            
            school_dict['is_online'] = is_online
            schools.append(school_dict)
        
        print(f"\n=== END DEBUG ===\n")
    
    stats = get_developer_stats()
    return render_template('developer_dashboard.html', 
                          schools=schools, 
                          stats=stats, 
                          subscription_plans=SUBSCRIPTION_PLANS,
                          session=session)

@app.route('/dev/add-school', methods=['POST'])
@require_developer
def dev_add_school():
    """Add new school with 7-day trial and initialize budget"""
    school_name = request.form.get('school_name', '').strip()
    username = request.form.get('username', '').strip()
    password = request.form.get('password', '').strip()
    financial_year = request.form.get('financial_year', '2026-2027')
    
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            password_hash = hash_password(password)
            
            # Use Starter Trial Plan by default
            trial_plan = get_plan('trial')
            trial_start = datetime.now()
            trial_end = trial_start + timedelta(days=trial_plan['duration_days'])
            
            cursor.execute('''INSERT INTO schools 
                (school_name, username, password_hash, is_active, subscription_type, subscription_status, 
                 subscription_start, subscription_end, expiry_date, must_change_password) 
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                (school_name, username, password_hash, 1, trial_plan['name'], 'TRIAL', 
                 trial_start.strftime('%Y-%m-%d'), trial_end.strftime('%Y-%m-%d'), trial_end.strftime('%Y-%m-%d'), 1))
            school_id = cursor.lastrowid
            
            # Initialize school settings with zero values
            # Using Term 1 as default for new schools
            default_term = "Term 1"
            cursor.execute('''INSERT INTO school_settings
                (school_id, academic_year, term, financial_year, school_name, total_grant)
                VALUES (?, ?, ?, ?, ?, ?)''',
                (school_id, financial_year, default_term, f"{financial_year}_{default_term}", school_name, 0))
            
            # Initialize budget allocation - insert all 42 rows from master template
            budget_structure = generate_budget_structure()
            for item in budget_structure:
                cursor.execute('''INSERT INTO budget_items
                    (school_id, academic_year, term, financial_year, template_row_id, item_key, pow_no, pow_name, sub_activity,
                     sub_item_description, code, total_allocation, monthly_allocations)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (school_id, financial_year, default_term, f"{financial_year}_{default_term}", item['template_row_id'], item['id'],
                     item['powNo'], item['powName'], item['subActivity'],
                     item['subItemDescription'], item['code'],
                     0, json.dumps(item.get('monthlyAllocations', {}))))
            
            try:
                log_action('DEVELOPER', session['user']['username'], 
                          f'ADD_SCHOOL: {school_name} (30-day trial, must change password)', 
                          school_id, request.remote_addr)
            except Exception as log_error:
                print(f"Error logging action: {log_error}")
    except Exception as e:
        print(f"Error adding school: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    
    return redirect(url_for('dev_dashboard'))

@app.route('/dev/send-expiry-warnings', methods=['POST'])
@require_developer
def dev_send_expiry_warnings():
    """Manually trigger expiry warnings for schools expiring in 14 days"""
    try:
        warned_count = auto_expiry_check()
        log_action('DEVELOPER', session['user']['username'], 
                  f'SEND_EXPIRY_WARNINGS: {warned_count} schools warned', 
                  None, request.remote_addr)
        return jsonify({'success': True, 'warned_count': warned_count})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/dev/reset-password/<int:school_id>', methods=['POST'])
@require_developer
def dev_reset_password(school_id):
    """Generate OTP for school"""
    otp = generate_otp()
    expires_at = datetime.now() + timedelta(hours=24)
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO password_reset_tokens (school_id, otp_code, expires_at) VALUES (?, ?, ?)', (school_id, otp, expires_at))
        log_action('DEVELOPER', session['user']['username'], 'RESET_PASSWORD', school_id, request.remote_addr)
    return jsonify({'success': True, 'otp': otp})

@app.route('/dev/lock-school/<int:school_id>', methods=['POST'])
@require_developer
def dev_lock_school(school_id):
    """Lock school account"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE schools SET is_locked = 1 WHERE id = ?', (school_id,))
        log_action('DEVELOPER', session['user']['username'], 'LOCK_SCHOOL', school_id, request.remote_addr)
    return jsonify({'success': True})

@app.route('/dev/unlock-school/<int:school_id>', methods=['POST'])
@require_developer
def dev_unlock_school(school_id):
    """Unlock school account"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('UPDATE schools SET is_locked = 0 WHERE id = ?', (school_id,))
        log_action('DEVELOPER', session['user']['username'], 'UNLOCK_SCHOOL', school_id, request.remote_addr)
    return jsonify({'success': True})

@app.route('/dev/send-message/<int:school_id>', methods=['POST'])
@require_developer
def dev_send_message(school_id):
    """Send message to school"""
    data = request.get_json()
    message = data.get('message')
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('INSERT INTO subscription_messages (school_id, message, message_type) VALUES (?, ?, ?)', (school_id, message, 'INFO'))
        log_action('DEVELOPER', session['user']['username'], 'SEND_MESSAGE', school_id, request.remote_addr)
    return jsonify({'success': True})

@app.route('/dev/delete-school/<int:school_id>', methods=['POST'])
@require_developer
def dev_delete_school(school_id):
    """Delete school and all associated data"""
    try:
        with get_db() as conn:
            cursor = conn.cursor()
            
            # Get school name for logging
            cursor.execute('SELECT school_name FROM schools WHERE id = ?', (school_id,))
            school = cursor.fetchone()
            school_name = school['school_name'] if school else 'Unknown'
            
            # Delete all school data
            cursor.execute('DELETE FROM school_settings WHERE school_id = ?', (school_id,))
            cursor.execute('DELETE FROM budget_items WHERE school_id = ?', (school_id,))
            cursor.execute('DELETE FROM credits WHERE school_id = ?', (school_id,))
            cursor.execute('DELETE FROM debits WHERE school_id = ?', (school_id,))
            cursor.execute('DELETE FROM school_sessions WHERE school_id = ?', (school_id,))
            cursor.execute('DELETE FROM subscription_messages WHERE school_id = ?', (school_id,))
            cursor.execute('DELETE FROM password_reset_tokens WHERE school_id = ?', (school_id,))
            cursor.execute('DELETE FROM schools WHERE id = ?', (school_id,))
            
            log_action('DEVELOPER', session['user']['username'], f'DELETE_SCHOOL: {school_name}', school_id, request.remote_addr)
        
        return jsonify({'success': True, 'message': f'School {school_name} deleted successfully'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/dev/update-subscription/<int:school_id>', methods=['POST'])
@require_developer
def dev_update_subscription(school_id):
    """Update school subscription using standardized plans"""
    try:
        data = request.get_json()
        plan_key = data.get('plan_key')
        
        if plan_key not in SUBSCRIPTION_PLANS:
            return jsonify({'success': False, 'error': 'Invalid subscription plan'}), 400
        
        plan = get_plan(plan_key)
        
        # Calculate expiry date
        start_date = datetime.utcnow().date()
        expiry_date = start_date + timedelta(days=plan['duration_days'])
        
        # Update school subscription in single transaction
        with get_db() as conn:
            cursor = conn.cursor()
            cursor.execute('''UPDATE schools 
                SET subscription_type = ?,
                    subscription_status = 'PAID',
                    subscription_start = ?,
                    subscription_end = ?,
                    expiry_date = ?
                WHERE id = ?''',
                (plan['name'], start_date.strftime('%Y-%m-%d'), 
                 expiry_date.strftime('%Y-%m-%d'), expiry_date.strftime('%Y-%m-%d'), school_id))
            
            log_action('DEVELOPER', session['user']['username'], 
                      f'UPDATE_SUBSCRIPTION: {plan["name"]} ({plan["duration_days"]} days, MWK {plan["price"]:,.2f})', 
                      school_id, request.remote_addr)
        
        return jsonify({
            'success': True, 
            'plan_name': plan['name'],
            'expiry_date': expiry_date.strftime('%Y-%m-%d'),
            'price': plan['price']
        })
    except Exception as e:
        print(f"Error updating subscription: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/grant-summary')
@require_login
def grant_summary():
    """Grant Summary dashboard"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    credits = get_credits(academic_year, term) or []
    debits = get_debits(academic_year, term) or []
    spending = calculate_spending(budget, debits)
    
    # Get total grant from settings instead of budget
    total_grant = get_total_grant(academic_year, term)
    
    total_budgeted = 0.0
    if budget and budget.get('items'):
        for item in budget.get('items', []):
            if isinstance(item, dict):
                alloc = item.get('totalAllocation', 0)
                if alloc is not None:
                    try:
                        total_budgeted += float(alloc)
                    except (ValueError, TypeError):
                        pass
    
    total_spent = 0.0
    for d in debits:
        if isinstance(d, dict):
            amt = d.get('amount', 0)
            if amt is not None:
                try:
                    total_spent += float(amt)
                except (ValueError, TypeError):
                    pass
    
    settings = get_settings()
    total_credited = float(settings.get('balance_bf', settings.get('balanceBF', 0)))
    for c in credits:
        if isinstance(c, dict):
            line_items = c.get('lineItems', [])
            if line_items:
                for li in line_items:
                    if isinstance(li, dict):
                        amt = li.get('amount', 0)
                        if amt is not None:
                            try:
                                total_credited += float(amt)
                            except (ValueError, TypeError):
                                pass
    
    # Monthly data for charts
    months = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]
    monthly_credits = [0] * 12
    monthly_debits = [0] * 12
    
    month_map = {"April": 0, "May": 1, "June": 2, "July": 3, "August": 4, "September": 5,
                 "October": 6, "November": 7, "December": 8, "January": 9, "February": 10, "March": 11}
    
    for c in credits:
        if isinstance(c, dict):
            month = c.get('month')
            if month in month_map:
                for li in c.get('lineItems', []):
                    if isinstance(li, dict):
                        amt = li.get('amount', 0)
                        if amt:
                            try:
                                monthly_credits[month_map[month]] += float(amt)
                            except (ValueError, TypeError):
                                pass
    
    for d in debits:
        if isinstance(d, dict):
            month = d.get('month')
            if month in month_map:
                amt = d.get('amount', 0)
                if amt:
                    try:
                        monthly_debits[month_map[month]] += float(amt)
                    except (ValueError, TypeError):
                        pass
    
    # Recent transactions for charts
    recent_credits_data = []
    for c in credits[-10:]:
        if isinstance(c, dict):
            total = sum(float(li.get('amount', 0)) for li in c.get('lineItems', []) if isinstance(li, dict))
            recent_credits_data.append({'date': c.get('date', ''), 'amount': total})
    
    recent_debits_data = []
    for d in debits[-10:]:
        if isinstance(d, dict):
            recent_debits_data.append({'date': d.get('date', ''), 'amount': float(d.get('amount', 0))})
    
    return render_template('grant_summary.html', 
                      budget=budget,
                      credits=credits,
                      debits=debits,
                      spending=spending,
                      total_grant=total_grant,
                      total_budgeted=total_budgeted,
                      total_spent=total_spent,
                      total_credited=total_credited,
                      academic_year=academic_year,
                      term=term,
                      settings=get_settings(),
                      current_page='dashboard',
                      months=months,
                      monthly_credits=monthly_credits,
                      monthly_debits=monthly_debits,
                      recent_credits_data=recent_credits_data,
                      recent_debits_data=recent_debits_data)

@app.route('/budget')
@require_login
def budget():
    """Budget allocation page - no self-healing, template is immutable"""
    school_id = get_current_school_id()
    academic_year = get_academic_year()
    term = get_term()
    
    budget = get_budget(academic_year, term)
    credits = get_credits(academic_year, term) or []
    debits = get_debits(academic_year, term) or []
    total_grant = get_total_grant(academic_year, term)
    settings = get_settings()
    
    return render_template('budget.html',
                      budget=budget,
                      credits=credits,
                      debits=debits,
                      total_grant=total_grant,
                      academic_year=academic_year,
                      term=term,
                      settings=settings,
                      current_page='budget')

@app.route('/export_budget_excel')
@require_login
def export_budget_excel():
    """Export Budget Allocation to Excel"""
    # Runtime check for openpyxl
    try:
        import openpyxl
    except ImportError:
        return "Excel export not available. Please install openpyxl: pip install openpyxl==3.1.2", 500
    
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    settings = get_settings()
    
    if not budget:
        return "No budget found", 404
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Allocation"
    
    # Styles
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Add centered logo at top
    try:
        logo_path = os.path.join('static', 'images', 'Malawi Government logo.png')
        if os.path.exists(logo_path):
            img = XLImage(logo_path)
            img.height = 60
            img.width = 60
            # Center the logo (column J is approximately center for 20 columns)
            ws.add_image(img, 'J1')
    except Exception as e:
        print(f"Could not add logo: {e}")
    
    # Title (below logo)
    ws.merge_cells('A4:T4')
    ws['A4'] = 'BUDGET ALLOCATION'
    ws['A4'].font = Font(bold=True, size=14)
    ws['A4'].alignment = center_align
    
    ws.merge_cells('A5:T5')
    ws['A5'] = f"{settings.get('ministry', 'NANJATI CDSS')} | {academic_year} {term}"
    ws['A5'].font = Font(size=10)
    ws['A5'].alignment = center_align
    
    # Headers (row 7)
    months = ["April", "May", "June", "July", "August", "September", 
              "October", "November", "December", "January", "February", "March"]
    
    headers = ['POW No', 'POW Name', 'Sub Activity', 'Sub Item Description', 'Code', 'Annual Allocation'] + months + ['Total Budgeted', 'Balance']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = center_align
    
    # Data rows (starting from row 8)
    row_num = 8
    for item in budget.get('items', []):
        ws.cell(row=row_num, column=1, value=item.get('powNo'))
        ws.cell(row=row_num, column=2, value=item.get('powName'))
        ws.cell(row=row_num, column=3, value=item.get('subActivity'))
        ws.cell(row=row_num, column=4, value=item.get('subItemDescription'))
        ws.cell(row=row_num, column=5, value=item.get('code'))
        ws.cell(row=row_num, column=6, value=item.get('totalAllocation', 0))
        
        # Monthly allocations
        monthly_allocs = item.get('monthlyAllocations', {})
        col_num = 7
        total_budgeted = 0
        for month in months:
            value = monthly_allocs.get(month, 0)
            ws.cell(row=row_num, column=col_num, value=value)
            total_budgeted += value
            col_num += 1
        
        # Total Budgeted
        ws.cell(row=row_num, column=19, value=total_budgeted)
        
        # Balance
        balance = item.get('totalAllocation', 0) - total_budgeted
        ws.cell(row=row_num, column=20, value=balance)
        
        # Apply borders and alignment
        for col in range(1, 21):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            if col >= 6:  # Numeric columns
                cell.alignment = Alignment(horizontal='right')
        
        row_num += 1
    
    # Grand totals
    ws.cell(row=row_num, column=4, value='GRAND TOTALS (POW 1-16):')
    ws.cell(row=row_num, column=4).font = Font(bold=True)
    
    # Calculate totals
    for col in range(6, 21):
        col_letter = openpyxl.utils.get_column_letter(col)
        formula = f"=SUM({col_letter}8:{col_letter}{row_num-1})"
        cell = ws.cell(row=row_num, column=col, value=formula)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        cell.border = border
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    for col in range(7, 19):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10
    ws.column_dimensions['S'].width = 12
    ws.column_dimensions['T'].width = 12
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Budget_Allocation_{academic_year}_{term}.xlsx".replace(' ', '_')
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/initialize_budget', methods=['GET', 'POST'])
@require_login
def initialize_budget():
    """Initialize budget - only creates template if no budget exists"""
    school_id = get_current_school_id()
    academic_year = get_academic_year()
    term = get_term()
    
    # Check if budget already exists
    existing_budget = get_budget(academic_year, term)
    if existing_budget and existing_budget.get('items'):
        # Budget already exists, return it
        return jsonify({'success': True, 'budget': existing_budget, 'message': 'Budget already exists'})
    
    if request.method == 'POST':
        data = request.get_json()
        budget_data = {
            'academicYear': data.get('academicYear', academic_year),
            'term': data.get('term', term),
            'schoolName': data.get('schoolName'),
            'totalGrant': data.get('totalGrant'),
            'items': generate_budget_structure(),
            'createdAt': datetime.now().isoformat()
        }
        
        if save_budget(budget_data):
            return jsonify({'success': True, 'budget': budget_data, 'message': 'Budget initialized successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to save budget'})
    
    return render_template('initialize_budget.html', academic_year=academic_year, term=term)

@app.route('/test_budget_save', methods=['GET'])
@require_login
def test_budget_save():
    """Test endpoint to verify budget save works"""
    school_id = get_current_school_id()
    academic_year = get_academic_year()
    term = get_term()
    
    # Create test budget data
    budget = get_budget(academic_year, term)
    if not budget or not budget.get('items'):
        return jsonify({'success': False, 'error': 'No budget found'})
    
    # Modify first 3 items
    for i in range(min(3, len(budget['items']))):
        budget['items'][i]['totalAllocation'] = 10000.0 * (i + 1)
        budget['items'][i]['monthlyAllocations']['April'] = 1000.0 * (i + 1)
    
    # Save
    result = save_budget(budget)
    
    # Verify
    budget_check = get_budget(academic_year, term)
    saved_values = [(item.get('totalAllocation'), item.get('monthlyAllocations', {}).get('April')) 
                    for item in budget_check.get('items', [])[:3]]
    
    return jsonify({
        'success': result,
        'saved_values': saved_values,
        'message': 'Test complete - check if values match [10000, 20000, 30000]'
    })

@app.route('/update_budget', methods=['POST'])
@require_login
def update_budget():
    """Update budget"""
    try:
        data = request.get_json()
        print(f"\n=== UPDATE BUDGET DEBUG ===")
        print(f"Received data keys: {data.keys() if data else 'None'}")
        print(f"Items count: {len(data.get('items', [])) if data else 0}")
        
        academic_year = data.get('academicYear', get_academic_year())
        term = data.get('term', get_term())
        print(f"Academic year: {academic_year}, Term: {term}")
        
        budget = get_budget(academic_year, term)
        if not budget:
            print("ERROR: Budget not found")
            return jsonify({'success': False, 'error': 'Budget not found for this financial year'})
        
        print(f"Existing budget items: {len(budget.get('items', []))}")
        
        # Update items with template_row_id preserved
        if 'items' in data:
            for i, new_item in enumerate(data['items']):
                if i < len(budget['items']):
                    # Preserve template_row_id from existing budget
                    new_item['template_row_id'] = budget['items'][i].get('template_row_id')
                    if i < 3:  # Log first 3 items
                        print(f"Item {i}: template_row_id={new_item.get('template_row_id')}, allocation={new_item.get('totalAllocation')}")
            budget['items'] = data['items']
        
        # Update settings if present
        if 'schoolName' in data or 'totalGrant' in data or 'balanceBF' in data:
            current_settings = get_settings()
            updated_settings = {
                'academicYear': academic_year,
                'term': term,
                'schoolName': data.get('schoolName', current_settings.get('schoolName')),
                'totalGrant': data.get('totalGrant', current_settings.get('totalGrant')),
                'balanceBF': data.get('balanceBF', current_settings.get('balanceBF')),
                'schoolAddress': current_settings.get('schoolAddress'),
                'ministry': current_settings.get('ministry'),
                'compiledBy': current_settings.get('compiledBy'),
                'enteredBy': current_settings.get('enteredBy'),
                'authorizingOfficer': current_settings.get('authorizingOfficer'),
                'authorizingAppointment': current_settings.get('authorizingAppointment'),
                'counterSign': current_settings.get('counterSign'),
                'counterAppointment': current_settings.get('counterAppointment')
            }
            save_settings(updated_settings)
            print(f"Updated settings for {updated_settings['schoolName']}")
            
            # Update session school name if changed
            if 'schoolName' in data:
                session['school_name'] = data['schoolName']
            
            # Ensure the budget object returned has the latest grant info
            budget['totalGrant'] = float(data.get('totalGrant', current_settings.get('totalGrant'))) + \
                                  float(data.get('balanceBF', current_settings.get('balanceBF')))
            budget['schoolName'] = data.get('schoolName', current_settings.get('schoolName'))
        
        budget['updatedAt'] = datetime.now().isoformat()
        
        print(f"Calling save_budget for {len(budget.get('items', []))} items...")
        result = save_budget(budget)
        print(f"Save result: {result}")
        
        if result:
            print("Budget update successful")
            return jsonify({'success': True, 'budget': budget, 'message': 'Budget updated successfully'})
        else:
            print("ERROR: save_budget returned False")
            return jsonify({'success': False, 'error': 'Failed to save budget items to database'})
    except Exception as e:
        print(f"\nERROR in update_budget: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/credits')
@app.route('/creditregister')
@app.route('/credit-register')
@require_login
def credits():
    """Credit register page"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    credits = get_credits(academic_year, term) or []
    debits = get_debits(academic_year, term) or []
    total_grant = get_total_grant(academic_year, term)
    settings = get_settings()
    
    # Calculate total credits amount
    total_credits_amount = 0.0
    for c in credits:
        if isinstance(c, dict):
            line_items = c.get('lineItems', [])
            for item in line_items:
                if isinstance(item, dict):
                    amt = item.get('amount', 0)
                    if amt:
                        try:
                            total_credits_amount += float(amt)
                        except (ValueError, TypeError):
                            pass
    
    return render_template('credits.html',
                      budget=budget,
                      credits=credits,
                      debits=debits,
                      total_credits_amount=total_credits_amount,
                      total_grant=total_grant,
                      academic_year=academic_year,
                      term=term,
                      settings=settings,
                      current_page='credits')

@app.route('/add_credit', methods=['POST'])
@require_login
def add_credit():
    """Add credit entry"""
    data = request.get_json()
    credit_data = {
        'id': f"credit_{datetime.now().timestamp()}",
        'date': data.get('date'),
        'month': data.get('month'),
        'lineItems': data.get('lineItems', []),
        'remarks': data.get('remarks'),
        'academicYear': data.get('academicYear', get_academic_year()),
        'term': data.get('term', get_term()),
        'createdAt': datetime.now().isoformat()
    }
    
    if save_credit(credit_data):
        return jsonify({'success': True, 'credit': credit_data})
    else:
        return jsonify({'success': False, 'error': 'Failed to save credit'})

@app.route('/delete_credit/<credit_id>', methods=['POST'])
@require_login
def delete_credit(credit_id):
    """Delete a credit entry"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    delete_school_credit(school_id, credit_id)
    return jsonify({'success': True})

@app.route('/debits')
@require_login
def debits():
    """Debit register page"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    credits = get_credits(academic_year, term)
    debits = get_debits(academic_year, term)
    available_funds = get_available_funds(academic_year, term)
    settings = get_settings()
    
    # Calculate grand totals for display
    total_spent = sum(d.get('amount', 0) for d in debits)
    total_credited = sum(
        sum(li.get('amount', 0) for li in c.get('lineItems', []))
        for c in credits
    )
    
    return render_template('debits.html',
                      budget=budget,
                      credits=credits,
                      debits=debits,
                      available_funds=available_funds,
                      total_spent=total_spent,
                      total_credited=total_credited,
                      academic_year=academic_year,
                      term=term,
                      settings=settings,
                      current_page='debits')

@app.route('/add_debit', methods=['POST'])
@require_login
def add_debit():
    """Add debit entry with unified transaction number"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    data = request.get_json()
    academic_year = data.get('academicYear', get_academic_year())
    term = data.get('term', get_term())
    
    # Generate unified transaction number (per term)
    transaction_number = generate_transaction_number(school_id, academic_year, term)
    
    debit_data = {
        'id': f"debit_{datetime.now().timestamp()}",
        'documentNumber': transaction_number,
        'date': data.get('date'),
        'month': data.get('month'),
        'itemId': data.get('itemId'),
        'subItemDescription': data.get('subItemDescription'),
        'code': data.get('code'),
        'description': data.get('description'),
        'amount': data.get('amount'),
        'amountWords': data.get('amountWords', ''),
        'supplierName': data.get('supplierName'),
        'position': data.get('position'),
        'chequeNumber': data.get('chequeNumber', ''),
        'academicYear': academic_year,
        'term': term,
        'createdAt': datetime.now().isoformat()
    }
    
    if save_debit(debit_data):
        return jsonify({'success': True, 'debit': debit_data})
    else:
        return jsonify({'success': False, 'error': 'Failed to save debit'})

@app.route('/delete_debit/<debit_id>', methods=['POST'])
@require_login
def delete_debit(debit_id):
    """Delete a debit entry"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    delete_school_debit(school_id, debit_id)
    return jsonify({'success': True})

@app.route('/loose_minute/<debit_id>')
@require_login
def loose_minute(debit_id):
    """Generate Loose Minute View - uses unified document number"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    debits = get_debits(academic_year, term)
    debit = next((d for d in debits if d.get('id') == debit_id), None)
    
    if not debit:
        return "Debit not found", 404
        
    return render_template('loose_minute.html', debit=debit, budget=budget)

@app.route('/gp10/<debit_id>')
@require_login
def gp10(debit_id):
    """Generate GP10 Voucher View"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    debits = get_debits(academic_year, term)
    debit = next((d for d in debits if d.get('id') == debit_id), None)
    
    if not debit:
        return "Debit not found", 404
        
    return render_template('gp10.html', debit=debit, budget=budget, settings=get_settings())

@app.route('/receipt/<debit_id>')
@require_login
def payment_receipt(debit_id):
    """Generate Payment Receipt View - uses unified document number"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    debits = get_debits(academic_year, term)
    debit = next((d for d in debits if d.get('id') == debit_id), None)
    
    if not debit:
        return "Debit not found", 404
        
    return render_template('receipt.html', debit=debit, budget=budget)

@app.route('/persist_ipdc/<debit_id>', methods=['POST'])
@require_login
def persist_ipdc_data(debit_id):
    """Save IPDC meeting details for a specific debit"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    ipdc_data = request.json
    success = update_debit_ipdc_fields(school_id, debit_id, ipdc_data)
    
    if success:
        return jsonify({'success': True})
    return jsonify({'success': False, 'error': 'Failed to save IPDC details'}), 500

@app.route('/get_next_ipdc_number')
@require_login
def get_next_ipdc_number_route():
    """Get next sequential IPDC minute number"""
    school_id = get_current_school_id()
    if not school_id:
        return jsonify({'success': False, 'error': 'Not authenticated'}), 401
    
    academic_year = get_academic_year()
    term = get_term()
    number = generate_ipdc_number(school_id, academic_year, term)
    return jsonify({'success': True, 'number': number})

@app.route('/ipdc_minute/<debit_id>')
@require_login
def ipdc_minute(debit_id):
    """Generate IPDC Minute View"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    debits = get_debits(academic_year, term)
    debit = next((d for d in debits if d.get('id') == debit_id), None)
    
    if not debit:
        return "Debit not found", 404
        
    return render_template('ipdc_minute.html', debit=debit, budget=budget, settings=get_settings())

@app.route('/save_combined_ipdc_data', methods=['POST'])
@require_login
def save_combined_ipdc_data():
    data = request.json
    session_id = str(uuid.uuid4())
    ipdc_sessions[session_id] = data
    return jsonify({'success': True, 'session_id': session_id})


@app.route('/combined_ipdc')
@require_login
def combined_ipdc():
    """Generate Combined IPDC Minute for multiple debits"""
    session_id = request.args.get('session_id')
    
    if session_id and session_id in ipdc_sessions:
        data = ipdc_sessions[session_id]
        meeting = data['meeting']
        items_data = data['items']
        
        academic_year = get_academic_year()
        term = get_term()
        debits = get_debits(academic_year, term)
        
        selected_items = []
        for item in items_data:
            debit = next((d for d in debits if str(d.get('id')) == str(item['debit_id'])), None)
            if debit:
                selected_items.append({
                    'debit': debit,
                    'quotes': item['quotes'],
                    'reason': item['reason']
                })
        
        meeting_info = {
            **meeting,
            'academicYear': academic_year,
            'term': term,
            'date': selected_items[0]['debit'].get('date', '') if selected_items else ''
        }
        
        return render_template('combined_ipdc_minute.html',
                          items=selected_items,
                          meeting=meeting_info,
                          settings=get_settings())
    
    # Fallback for old style URLs (optional)
    debit_ids = request.args.getlist('ids')
    academic_year = get_academic_year()
    term = get_term()
    debits = get_debits(academic_year, term)
    selected_debits = [d for d in debits if str(d.get('id')) in debit_ids]
    
    items = [{'debit': d, 'quotes': [{'name': d['supplierName'], 'amount': d['amount'], 'is_selected': True}], 'reason': 'Selected after review'} for d in selected_debits]
    
    meeting_info = {
        'venue': request.args.get('venue', ''),
        'minute_no': request.args.get('minute_no', ''),
        'members': request.args.get('members', ''),
        'opening_prayer': request.args.get('opening_prayer', ''),
        'closing_prayer': request.args.get('closing_prayer', ''),
        'date': selected_debits[0].get('date', '') if selected_debits else '',
        'academicYear': academic_year,
        'term': term
    }
    
    return render_template('combined_ipdc_minute.html',
                      items=items,
                      meeting=meeting_info,
                      settings=get_settings())


@app.route('/settings', methods=['GET', 'POST'])
@require_login
def settings():
    """System settings page with subscription info"""
    school_id = get_current_school_id()
    
    if request.method == 'POST':
        form_type = request.form.get('form_type')
        
        if form_type == 'financial_year':
            # Update academic year and term in session and settings
            academic_year = request.form.get('academicYear', '').strip()
            term = request.form.get('term', '').strip()
            
            # Sanitize: replace en-dash, em-dash with hyphen, remove spaces
            academic_year = academic_year.replace('–', '-').replace('—', '-').replace(' ', '')
            
            # Save to database for persistence
            if school_id:
                save_term_settings(school_id, academic_year, term)
                session['academic_year'] = academic_year
                session['term'] = term
                
                # Initialize budget structure for new term if needed
                initialize_budget_for_year(school_id, academic_year, term)
                
            flash(f"System switched to {academic_year} {term}", "success")
            return redirect(url_for('grant_summary'))
        else:
            # Update all settings including school address and total grant
            financial_year = get_financial_year()
            
            settings_data = {
                'compiledBy': request.form.get('compiledBy', ''),
                'enteredBy': request.form.get('enteredBy', ''),
                'authorizingOfficer': request.form.get('authorizingOfficer', ''),
                'authorizingAppointment': request.form.get('authorizingAppointment', ''),
                'counterSign': request.form.get('counterSign', ''),
                'counterAppointment': request.form.get('counterAppointment', ''),
                'ministry': request.form.get('ministry', 'Education'),
                'schoolName': request.form.get('schoolName', ''),
                'schoolAddress': request.form.get('schoolAddress', ''),
                'financialYear': financial_year,
                'totalGrant': float(request.form.get('totalGrant', 0)),
                'balanceBF': float(request.form.get('balanceBF', 0))
            }
            save_settings(settings_data)
            return redirect(url_for('settings', success=True))
    
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    credits = get_credits(academic_year, term) or []
    debits = get_debits(academic_year, term) or []
    total_grant = get_total_grant(academic_year, term)
    
    # Get subscription info
    subscription_info = get_subscription_info(school_id) if school_id else None
    
    return render_template('settings.html', 
                          settings=get_settings(),
                          budget=budget,
                          credits=credits,
                          debits=debits,
                          academic_year=academic_year,
                          term=term,
                          total_grant=total_grant,
                          subscription_info=subscription_info,
                          current_page='settings',
                          success=request.args.get('success'))

@app.route('/itemized')
@require_login
def itemized():
    """Itemized spending page"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    debits = get_debits(academic_year, term) or []
    
    if term == 'Term 1':
        term_months = ['September', 'October', 'November', 'December']
    elif term == 'Term 2':
        term_months = ['January', 'February', 'March']
    else:
        term_months = ['April', 'May', 'June', 'July']
        
    itemized_data = []
    monthly_totals = {m: 0 for m in term_months}
    total_alloc = 0
    total_used = 0
    
    if budget and budget.get('items'):
        for item in budget.get('items', []):
            try:
                item_alloc = float(item.get('totalAllocation', 0) or 0)
            except (ValueError, TypeError):
                item_alloc = 0.0
                
            item_used = 0.0
            month_amounts = {m: 0.0 for m in term_months}
            
            for d in debits:
                if d.get('itemId') == item.get('id'):
                    try:
                        amt = float(d.get('amount', 0) or 0)
                    except (ValueError, TypeError):
                        amt = 0.0
                    month = d.get('month')
                    if month in term_months:
                        month_amounts[month] += amt
                        monthly_totals[month] += amt
                    # We only sum debits that belong to this term
                    item_used += amt
            
            # Only include items that have a budget allocation
            if item_alloc > 0:
                itemized_data.append({
                    'subItemDescription': item.get('subItemDescription'),
                    'code': item.get('code'),
                    'totalAllocation': item_alloc,
                    'month_amounts': month_amounts,
                    'totalUsed': item_used,
                    'balance': item_alloc - item_used
                })
                total_alloc += item_alloc
                total_used += item_used

    totals = {
        'totalAllocation': total_alloc,
        'monthly_totals': monthly_totals,
        'totalUsed': total_used,
        'balance': total_alloc - total_used
    }

    return render_template('itemized.html',
                      itemized_data=itemized_data,
                      totals=totals,
                      term_months=term_months,
                      academic_year=academic_year,
                      term=term,
                      settings=get_settings(),
                      current_page='itemized')

@app.route('/tracking')
@require_login
def tracking():
    """Spending tracking page"""
    academic_year = get_academic_year()
    term = get_term()
    budget = get_budget(academic_year, term)
    credits = get_credits(academic_year, term) or []
    debits = get_debits(academic_year, term) or []
    spending = calculate_spending(budget, debits)
    total_grant = get_total_grant(academic_year, term)
    
    # Calculate total spent
    total_spent = 0.0
    for d in debits:
        if isinstance(d, dict):
            amt = d.get('amount', 0)
            if amt:
                try:
                    total_spent += float(amt)
                except (ValueError, TypeError):
                    pass
    
    return render_template('tracking.html',
                      budget=budget,
                      credits=credits,
                      debits=debits,
                      spending=spending,
                      total_spent=total_spent,
                      total_grant=total_grant,
                      academic_year=academic_year,
                      term=term,
                      settings=get_settings(),
                      current_page='tracking')

@app.route('/debug/available_funds')
@require_login
def debug_available_funds():
    """Debug endpoint to check credit-debit linking"""
    academic_year = get_academic_year()
    term = get_term()
    available = get_available_funds(academic_year, term)
    return jsonify({
        'academic_year': academic_year,
        'term': term,
        'total_items': len(available),
        'available_funds': {item_id: {'budgeted': f['budgeted'], 'credited': f['credited'], 'spent': f['spent'], 'balance': f['balance']} for item_id, f in available.items()}
    })

@app.route('/debug/budget_items')
@require_login
def debug_budget_items():
    """Debug endpoint to check budget items in database"""
    school_id = get_current_school_id()
    academic_year = get_academic_year()
    term = get_term()
    
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) as count FROM budget_items WHERE school_id = ? AND academic_year = ? AND term = ?',
                      (school_id, academic_year, term))
        count = cursor.fetchone()['count']
        
        cursor.execute('SELECT template_row_id, item_key, pow_no, sub_item_description FROM budget_items WHERE school_id = ? AND academic_year = ? AND term = ? LIMIT 5',
                      (school_id, academic_year, term))
        sample = [dict(row) for row in cursor.fetchall()]
    
    return jsonify({
        'school_id': school_id,
        'academic_year': academic_year,
        'term': term,
        'total_items': count,
        'sample_items': sample,
        'expected': 42
    })

@app.route('/debug/activity')
@require_developer
def debug_activity():
    """Debug endpoint to check activity tracking"""
    with get_db() as conn:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, school_name, last_activity, last_login,
                   datetime('now') as current_time
            FROM schools 
            WHERE school_name != 'DEVELOPER_ACCOUNT'
            ORDER BY last_activity DESC
        ''')
        schools = [dict(row) for row in cursor.fetchall()]
        
        # Calculate time differences
        for school in schools:
            if school['last_activity']:
                try:
                    last_time = datetime.strptime(school['last_activity'], '%Y-%m-%d %H:%M:%S')
                    diff = datetime.utcnow() - last_time
                    school['minutes_ago'] = diff.total_seconds() / 60
                    school['is_online'] = diff.total_seconds() < 300  # 5 minutes
                except:
                    school['minutes_ago'] = None
                    school['is_online'] = False
            else:
                school['minutes_ago'] = None
                school['is_online'] = False
    
    return jsonify({
        'current_utc': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
        'schools': schools
    })

if __name__ == '__main__':
    print("🚀 Starting Grant Management System - Multi-Tenant")
    print("📱 Access at: http://localhost:5176")
    print("\n🔐 Default Login Credentials:")
    print("   School: admin / admin123")
    print("   Developer: Type 'devaccess' on login screen")
    print("   Developer Email: juniornsambe@yahoo.com")
    print("   Developer Password: blessings19831983/")
    
    # Register shutdown handler to checkpoint WAL
    import atexit
    def checkpoint_on_exit():
        try:
            import sqlite3
            conn = sqlite3.connect(DB_FILE)
            conn.execute('PRAGMA wal_checkpoint(TRUNCATE)')
            conn.close()
            print("\n✅ Database checkpointed on shutdown")
        except Exception as e:
            print(f"\n⚠️ Checkpoint error: {e}")
    
    atexit.register(checkpoint_on_exit)

    _reloader = os.environ.get("FLASK_USE_RELOADER", "1").strip().lower() not in ("0", "false", "no", "off")
    app.run(debug=True, host="0.0.0.0", port=5176, use_reloader=_reloader, threaded=True)
